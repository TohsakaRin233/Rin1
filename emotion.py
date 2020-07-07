from snownlp import SnowNLP
from openpyxl import load_workbook
excel_dir = 'C:\\Users\\Administrator\\Desktop\\result.xlsx'
excel = load_workbook(excel_dir)
table = excel.get_sheet_by_name('Sheet1')
nrows = table.max_row
ncols = table.max_column
count_good=0
count_bad=0
count_normal=0
for i in range(2,nrows):
    s =SnowNLP(table.cell(row=i,column=2).value)
    if(s.sentiments > 0.6):
        count_good+=1
    elif(s.sentiments <0.4):
        count_bad+=1
    else:
        count_normal+=1
    print(str(s.sentiments) +" "+ table.cell(row=i,column=2).value)
print("积极情绪：" + str(count_good/(count_good+count_normal+count_bad)))
print("中间情绪："+ str(count_normal/(count_good+count_normal+count_bad)))
print("消极情绪："+ str(count_bad/(count_good+count_normal+count_bad)))
print("总评论数："+ str(count_good+count_normal+count_bad))
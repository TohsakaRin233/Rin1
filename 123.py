from openpyxl import load_workbook
excel=load_workbook('C:\\Users\\Administrator\\Desktop\\result.xlsx')
table = excel.get_sheet_by_name('Sheet1')
nrows = table.max_row
ncols = table.max_column
for i in range(2,nrows):
    for j in range(4,ncols):
        if(table.cell(row=i,column=j).value == None):
            table.cell(row=i,column=j).value = 0
            print("success")
excel.save('C:\\Users\\Administrator\\Desktop\\result.xlsx')
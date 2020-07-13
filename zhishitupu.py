from openpyxl import load_workbook
excel=load_workbook('C:\\Users\\Administrator\\Desktop\\result.xlsx')
table = excel.get_sheet_by_name('Sheet1')
nrows = table.max_row
ncols = table.max_column
for i in range(4,ncols):
    if(int(table.cell(row=251,column=i).value) < 5):
        table.delete_cols(i)
        print("success")
        excel.save('C:\\Users\\Administrator\\Desktop\\result.xlsx')

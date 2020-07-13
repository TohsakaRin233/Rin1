import jieba
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from wordcloud import WordCloud
excel=load_workbook('C:\\Users\\Administrator\\Desktop\\result.xlsx')
table = excel.get_sheet_by_name('Sheet1')
nrows = table.max_row
ncols = table.max_column
mytext_list = []
nowcols = 4

def is_ustr(in_str):
    out_str=''
    for i in range(len(in_str)):
        if is_uchar(in_str[i]):
            out_str=out_str+in_str[i]
        else:
            out_str=out_str+' '
    return out_str
def is_uchar(uchar):
    if uchar >= u'\u4e00' and uchar<=u'\u9fa5':
        return True
    return False


for i in range(2,nrows):
    com =table.cell(row=i,column=2).value.replace(" ","").replace("\n","").replace("\r","")
    com = is_ustr(com).replace(" ","")
    segs = jieba.cut(com)
    for seg in segs:
        if len(seg) != 1:
            if seg not in mytext_list:
                mytext_list.append(seg)
                table.cell(row = 1,column = nowcols).value = seg
                table.cell(row=i, column=nowcols).value = 1
                nowcols+=1
            else:
                table.cell(row=i, column=mytext_list.index(seg)+4).value = 1
excel.save('C:\\Users\\Administrator\\Desktop\\result.xlsx')
print(mytext_list)


'''cloud_text = " ".join(mytext_list)
print(cloud_text)

# 制作词云
wordcloud = WordCloud(background_color="white", font_path="C:\\Windows\\Fonts\\msyh.ttc", height=300, width = 400).generate(cloud_text)

# 图片展示

plt.imshow(wordcloud)
plt.axis("off")
plt.show()

# 将词云图片导出到当前文件夹
#wordcloud.to_file("C:\\Users\\Administrator\\Desktop\\wordCloud.png")'''

